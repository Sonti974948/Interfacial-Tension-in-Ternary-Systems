
import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import streamlit as st


# In[2]:
st.set_page_config(layout="wide", page_title="Interfacial Tension AcA")
st.title("INTERFACIAL TENSION IN TERNARY SYSTEMS")
st.header("EXTRACTION OF ACETIC ACID FROM WATER")

st.write(r"""The interfacial tension calculations for Acetic Acid in Water and organic phase are done using the **LI-FU EQUATION**,where the ternary system interfacial tension is calculated using
biphasic interfacial tension, and the concentration of of the components in the system.  """)

#st.caption("Poling, B. E., Prausnitz, J. M., & O’connell, J. P. (2001). Properties of gases and liquids. McGraw-Hill Education.")

workbook = load_workbook(filename="CalcData.xlsx")
sheet1=workbook["Data"]


# In[3]:


#Arrays for data sets 
phi=np.zeros(3)
code=[1,3,4]
M_wt=np.zeros(3)
mu=np.zeros(3)
Tc=np.zeros(3)
Pc=np.zeros(3)
omega=np.zeros(3)
rho=np.zeros(3)
A=np.zeros(3)
B=np.zeros(3)
R=np.zeros(3)
Q=np.zeros(3)
aij=np.zeros(3)
aji=np.zeros(3)


with st.sidebar:
    name=st.selectbox("Choose the ORGANIC ENTRAINER in the ternary system",["NBA","IPA"])
    Texp=st.selectbox("Choose the temperature of system",[298,303,308])
    


if name=="NBA":
    code[0]=1
elif name=="IPA":
    code[0]=2
    

    
for i in range(3):
    if code[i]==4:
        mu_w=[0.889,0.797,0.719]
        
    
for i in range(3):
    k=code[i]
    phi[i]=sheet1.cell(row=k+11,column=5).value
    M_wt[i]=sheet1.cell(row=k+11,column=6).value
    mu[i]=sheet1.cell(row=k+11,column=7).value
    Tc[i]=sheet1.cell(row=k+11,column=8).value
    Pc[i]=sheet1.cell(row=k+11,column=9).value
    omega[i]=sheet1.cell(row=k+11,column=10).value
    rho[i]=sheet1.cell(row=k+11,column=11).value
    A[i]=sheet1.cell(row=k+11,column=12).value
    B[i]=sheet1.cell(row=k+11,column=13).value
    R[i]=sheet1.cell(row=k+11,column=14).value
    Q[i]=sheet1.cell(row=k+11,column=15).value
    aij[i]=sheet1.cell(row=k+11,column=16).value
    aji[i]=sheet1.cell(row=k+11,column=17).value


print(R)

# In[4]:


# In[5]:


org=code[0]

if org==1:
    sheet2=workbook["NBA"]
elif org==2:
    sheet2=workbook["IPA"]


# In[6]:

# In[10]:

if Texp==298:
    n=2 
elif Texp==303:
    n=8
elif Texp==308:
    n=14


conc_aq=np.zeros(3)
sigma=np.zeros(5)
phi_mix_aq=np.zeros(5)
Da_aq=np.zeros(5)
for i in range(5):
    conc_org1=sheet2.cell(row=i+n,column=12).value
    conc_aq1=sheet2.cell(row=i+n,column=14).value
    conc_org2=sheet2.cell(row=i+n,column=15).value
    conc_aq2=sheet2.cell(row=i+n,column=17).value
    conc_AcA=sheet2.cell(row=i+n,column=13).value
    # st.write(conc_aq2,conc_aq1,conc_org1,conc_org2)
    # r1=conc_org1*R[0]+conc_aq1*R[2]+(1-conc_org1-conc_aq1)*R[1]
    # r2=conc_org2*R[0]+conc_aq2*R[2]+(1-conc_org2-conc_aq2)*R[1]
    # phi1=(conc_aq1*R[2])/r1 
    # phi2=(conc_aq2*R[2])/r2 
    X=-np.log(conc_org1+conc_aq2+conc_AcA)
    k=0.467-0.185*X+0.016*np.power(X,2)
    W=(aij[0]+aji[0])/10*(1e7)
    if code[0]==1:
        surf=14.5
        X0=2.0959 #2.0959
    else:
        surf=12.446
        X0=2.393 #2.1597
    #st.write(X,k,surf)
    #sigma[i]=(3.14e-9)*(1-k)*W*np.power(phi1-phi2,2)
    sigma[i]=surf*np.power(X/X0,1-k)

    


sig_exp=np.zeros(5)
for i in range(5):
    sig_exp[i]=sheet2.cell(row=i+n,column=11).value/2

conc_AcAorg=np.zeros(5)
for i in range(5):
    conc_AcAorg[i]=sheet2.cell(row=i+n,column=16).value

conc_AcAaq=np.zeros(5)
for i in range(5):
    conc_AcAaq[i]=sheet2.cell(row=i+n,column=13).value

error=-(sig_exp-sigma)/sigma*100

if code[0]==1 and Texp==298:
            fn1='NBA_298_int.png'
            
elif code[0]==1 and Texp==303:
            fn1='NBA_303_int.png'
            
elif code[0]==1 and Texp==308:
            fn1='NBA_308_int.png'
            
elif code[0]==2 and Texp==298:
            fn1='IPA_298_int.png'
            
elif code[0]==2 and Texp==303:
            fn1='IPA_303_int.png'
            
elif code[0]==2 and Texp==308:
            fn1='IPA_308_int.png'
            

fig2,ax2=plt.subplots()
ax2.plot(conc_AcAaq,sig_exp,'ko'),
ax2.plot(conc_AcAaq,sigma,'k-')
ax2.legend(['GROMACS', "Li-Fu"], bbox_to_anchor=(1.8, 1), ncol=2)
plt.xlabel("Mole fraction of AcA in aqueous phase")
plt.ylabel("Interfacial Tension ($ mN \, m^{-1}$)")
plt.title('Interfacial Tension vs mol fraction of acetic acid',pad=2)



# with st.expander(label="Raw data of diffusivity in organic phase (error in %)"):
#     for i in range(5):
#         cols=st.columns(3)
#         cols[1].metric('GROMACS',np.around(Da_orgexp[i]*(1e5),decimals=3))
#         cols[2].metric(method,np.around(Da_org[i]*(1e5),decimals=3),np.around(error_org[i],decimals=2))
#         cols[0].metric('AcA in organic phase (mol/mol)',np.around(conc_AcAorg[i],decimals=3))

st.pyplot(fig2)

plt.savefig(fn1,bbox_inches='tight')
with open(fn1, "rb") as img:
    btn = st.download_button(
        label="Download as image",
        data=img,
        file_name=fn1,
        mime="image/png"
    )
with st.expander(label="Raw data of Interfacial Tension [mN/m] (error in %)"):
    for i in range(5):
        cols=st.columns(3)
        cols[1].metric('GROMACS',np.around(sig_exp[i],decimals=3))
        cols[2].metric("Li-Fu",np.around(sigma[i],decimals=3),np.around(error[i],decimals=2))
        cols[0].metric('AcA in aqueous phase (mol/mol)',np.around(conc_AcAaq[i],decimals=3))


# if Texp==298:
#     Da_aqp=wc.get_Da_w(Texp,mu[2],phi[2],M_wt[2])
#     Da_orgp=wc.get_Da_org(Texp,mu[0],phi[0],M_wt[0])

# else:
#     Da_aqp=wc.get_Da_w(Texp,muT[2],phi[2],M_wt[2])
#     Da_orgp=wc.get_Da_org(Texp,muT[0],phi[0],M_wt[0])

# print(Da_aqp)

st.header("Reference")
st.caption("Poling, B. E., Prausnitz, J. M., & O’connell, J. P. (2001). Properties of gases and liquids. McGraw-Hill Education.")

